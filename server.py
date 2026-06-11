"""
GoldBazaar Local Server
=======================
Run from the GoldBazaar folder:
    python server.py

Opens the site at:  http://localhost:8000

All form submissions are saved automatically to:
    customer_data/GoldBazaar_Leads.xlsx
      - Sheet "Contact Enquiries"  ← Get in Touch form
      - Sheet "Vendor Signups"     ← Vendor Onboarding (OTP + full details)
"""

import http.server
import json
import os
import threading
from datetime import datetime
from pathlib import Path

# ── Paths ──
BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "customer_data" / "GoldBazaar_Leads.xlsx"
PORT       = 8000

# ── openpyxl ──
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    raise

# Thread lock so concurrent requests don't corrupt the file
_lock = threading.Lock()


def now_str():
    return datetime.now().strftime("%d-%b-%Y, %I:%M %p")


def append_row(sheet_name: str, row: list):
    """Append one row to the given sheet in the Excel file."""
    with _lock:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        # Find next empty row (header is row 1, data starts row 2)
        next_row = ws.max_row + 1
        row_num  = next_row - 1          # sequential # column
        ws.cell(row=next_row, column=1, value=row_num)
        for col, val in enumerate(row, start=2):
            ws.cell(row=next_row, column=col, value=val)
        wb.save(EXCEL_PATH)


class GoldBazaarHandler(http.server.SimpleHTTPRequestHandler):
    """Serves static files + handles POST API endpoints."""

    def do_POST(self):
        length  = int(self.headers.get("Content-Length", 0))
        raw     = self.rfile.read(length)
        try:
            data = json.loads(raw)
        except Exception:
            self._respond(400, {"error": "Invalid JSON"})
            return

        if self.path == "/api/contact":
            self._handle_contact(data)
        elif self.path == "/api/vendor":
            self._handle_vendor(data)
        else:
            self._respond(404, {"error": "Unknown endpoint"})

    def _handle_contact(self, d):
        """Save a Get in Touch submission → Contact Enquiries sheet."""
        try:
            append_row("Contact Enquiries", [
                now_str(),
                d.get("name",    ""),
                d.get("mobile",  ""),
                d.get("type",    ""),
                d.get("city",    ""),
                d.get("message", ""),
            ])
            print(f"[Contact] {d.get('name')} — {d.get('mobile')}")
            self._respond(200, {"ok": True})
        except Exception as e:
            print(f"[Contact ERROR] {e}")
            self._respond(500, {"error": str(e)})

    def _handle_vendor(self, d):
        """Save a Vendor Signup → Vendor Signups sheet."""
        try:
            append_row("Vendor Signups", [
                now_str(),
                d.get("mobile",       ""),
                d.get("businessName", ""),
                d.get("ownerName",    ""),
                d.get("category",     ""),
                d.get("services",     ""),
                d.get("city",         ""),
                d.get("state",        ""),
                d.get("pincode",      ""),
                d.get("status",       "Application Submitted"),
            ])
            print(f"[Vendor] {d.get('businessName')} — {d.get('mobile')}")
            self._respond(200, {"ok": True})
        except Exception as e:
            print(f"[Vendor ERROR] {e}")
            self._respond(500, {"error": str(e)})

    def _respond(self, code, payload):
        body = json.dumps(payload).encode()
        self.send_response(code)
        self.send_header("Content-Type",  "application/json")
        self.send_header("Content-Length", len(body))
        # Allow requests from the same local origin
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        """Pre-flight CORS."""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def log_message(self, fmt, *args):
        # Suppress noisy static-file logs, keep API logs
        if "/api/" in (args[0] if args else ""):
            super().log_message(fmt, *args)


if __name__ == "__main__":
    os.chdir(BASE_DIR)          # serve files from the GoldBazaar folder
    server = http.server.HTTPServer(("", PORT), GoldBazaarHandler)
    print(f"\n  ✅  GoldBazaar running at  http://localhost:{PORT}")
    print(f"  📊  Excel file:  customer_data/GoldBazaar_Leads.xlsx")
    print(f"      Sheets:  'Contact Enquiries'  |  'Vendor Signups'")
    print(f"\n  Press Ctrl+C to stop.\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
